# views.py

from django.shortcuts import render, redirect, get_object_or_404
from django.contrib.auth import authenticate, login, logout
from django.contrib.auth.decorators import login_required
from django.contrib import messages
from django.http import HttpResponse, JsonResponse
from django.views.decorators.csrf import csrf_exempt
from django.db.models import Sum
from datetime import date, timedelta
import csv
import openpyxl
from openpyxl.utils import get_column_letter
from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer
from reportlab.lib.pagesizes import A4
from reportlab.lib import colors
from reportlab.lib.styles import getSampleStyleSheet
import json
import re
import logging

from django.conf import settings
import google.generativeai as genai
from pymongo import MongoClient

from .models import Expense, Category, UserProfile
from .forms import ExpenseForm, SimpleUserCreationForm, UserProfileForm

# Configure logging
logger = logging.getLogger(__name__)

# Configure Gemini AI
genai.configure(api_key=settings.GEMINI_API_KEY)

# ---------- MongoDB for AI Memory ----------
client = MongoClient("mongodb://localhost:27017/")  # or Atlas URI
db = client["ai_assistant"]
chat_collection = db["conversations"]
summary_collection = db["summaries"]  # new collection for long-term summaries


def save_message(user_id, role, content):
    """Save a chat message for a specific user"""
    chat_collection.insert_one({
        "user_id": user_id,
        "role": role,
        "content": content
    })


def load_history(user_id, limit=10):
    """Load last N messages for a specific user"""
    return list(chat_collection.find(
        {"user_id": user_id},
        {"_id": 0}
    ).sort("_id", -1).limit(limit))[::-1]  # reverse to keep order


def save_summary(user_id, summary):
    """Save or update long-term summary for a user"""
    summary_collection.update_one(
        {"user_id": user_id},
        {"$set": {"summary": summary}},
        upsert=True
    )


def load_summary(user_id):
    """Load summary memory for a user"""
    doc = summary_collection.find_one({"user_id": user_id}, {"_id": 0})
    return doc["summary"] if doc else ""


def summarize_old_chats(user_id):
    """Summarize old chats if messages exceed a threshold"""
    total_msgs = chat_collection.count_documents({"user_id": user_id})

    if total_msgs > 50:  # threshold
        old_msgs = list(chat_collection.find({"user_id": user_id}).sort("_id", 1))  # oldest first
        history_text = "\n".join([f"{m['role']}: {m['content']}" for m in old_msgs])

        # Ask Gemini to create a summary
        model = genai.GenerativeModel("gemini-1.5-flash")
        response = model.generate_content(
            f"Summarize the following conversation into key facts and user preferences:\n\n{history_text}"
        )
        summary = response.text.strip() if response and hasattr(response, "text") else ""

        if summary:
            save_summary(user_id, summary)

        # Delete old chats after summarizing
        chat_collection.delete_many({"user_id": user_id})


# ---------- User Auth ----------
def register(request):
    if request.method == "POST":
        form = SimpleUserCreationForm(request.POST)
        if form.is_valid():
            user = form.save()
            UserProfile.objects.create(user=user)  # create profile automatically
            messages.success(request, "Account created successfully!")
            return redirect("login")
    else:
        form = SimpleUserCreationForm()
    return render(request, "registration/register.html", {"form": form})


def user_login(request):
    if request.method == "POST":
        username = request.POST.get("username")
        password = request.POST.get("password")
        user = authenticate(request, username=username, password=password)
        if user:
            login(request, user)
            return redirect("home")
        messages.error(request, "Invalid credentials")
    return render(request, "registration/login.html")


def user_logout(request):
    logout(request)
    return redirect("login")


# ---------- Helper (Apply Filters) ----------
def filter_expenses(request, expenses):
    today = date.today()
    category = request.GET.get("category")
    date_range = request.GET.get("date_range")
    search = request.GET.get("search")
    sort_by = request.GET.get("sort_by")

    if category:
        expenses = expenses.filter(category__name=category)

    if date_range == "today":
        expenses = expenses.filter(date=today)
    elif date_range == "week":
        start_week = today - timedelta(days=7)
        expenses = expenses.filter(date__gte=start_week)
    elif date_range == "month":
        expenses = expenses.filter(date__month=today.month, date__year=today.year)

    if search and search.strip():
        expenses = expenses.filter(title__icontains=search)

    if sort_by == "newest":
        expenses = expenses.order_by("-date")
    elif sort_by == "oldest":
        expenses = expenses.order_by("date")
    elif sort_by == "high":
        expenses = expenses.order_by("-amount")
    elif sort_by == "low":
        expenses = expenses.order_by("amount")
    else:
        expenses = expenses.order_by("-date")  # default

    return expenses


# ---------- Dashboard ----------
@login_required
def home(request):
    expenses = Expense.objects.filter(user=request.user)
    expenses = filter_expenses(request, expenses)

    today = date.today()
    week_start = today - timedelta(days=today.weekday())

    total_today = expenses.filter(date=today).aggregate(Sum('amount'))['amount__sum'] or 0
    total_week = expenses.filter(date__gte=week_start).aggregate(Sum('amount'))['amount__sum'] or 0
    total_month = expenses.filter(date__month=today.month).aggregate(Sum('amount'))['amount__sum'] or 0
    total_expense = expenses.aggregate(Sum('amount'))['amount__sum'] or 0

    category_data_dict = expenses.values('category__name').annotate(total=Sum('amount'))
    category_labels = [c['category__name'] for c in category_data_dict]
    category_data = [c['total'] for c in category_data_dict]

    return render(request, "tracker/home.html", {
        "expenses": expenses,
        "categories": Category.objects.all(),
        "total_expense": total_expense,
        "total_today": total_today,
        "total_week": total_week,
        "total_month": total_month,
        "category_labels": category_labels,
        "category_data": category_data,
    })


# ---------- Expenses ----------
@login_required
def add_expense(request):
    if request.method == "POST":
        form = ExpenseForm(request.POST, request.FILES)
        if form.is_valid():
            expense = form.save(commit=False)
            expense.user = request.user
            expense.save()
            messages.success(request, "Expense added!")
            return redirect("home")
    else:
        form = ExpenseForm()
    return render(request, "tracker/add_expense.html", {"form": form})


@login_required
def edit_expense(request, expense_id):
    expense = get_object_or_404(Expense, id=expense_id, user=request.user)
    if request.method == "POST":
        form = ExpenseForm(request.POST, request.FILES, instance=expense)
        if form.is_valid():
            form.save()
            messages.success(request, "Expense updated!")
            return redirect("home")
    else:
        form = ExpenseForm(instance=expense)
    return render(request, "tracker/edit_expense.html", {"form": form})


@login_required
def delete_expense(request, expense_id):
    expense = get_object_or_404(Expense, id=expense_id, user=request.user)
    expense.delete()
    messages.success(request, "Expense deleted!")
    return redirect("home")


# ---------- Export CSV ----------
@login_required
def export_csv(request):
    expenses = filter_expenses(request, Expense.objects.filter(user=request.user))

    response = HttpResponse(content_type='text/csv')
    response['Content-Disposition'] = 'attachment; filename="expenses.csv"'
    writer = csv.writer(response)
    writer.writerow(['Title', 'Amount', 'Category', 'Date', 'Notes'])
    for exp in expenses:
        writer.writerow([
            exp.title,
            exp.amount,
            exp.category.name if exp.category else '',
            exp.date,
            exp.notes
        ])
    return response


# ---------- Export Excel ----------
@login_required
def export_excel(request):
    expenses = filter_expenses(request, Expense.objects.filter(user=request.user))

    workbook = openpyxl.Workbook()
    worksheet = workbook.active
    worksheet.title = "Expenses"
    headers = ["Title", "Amount", "Category", "Date", "Notes"]
    worksheet.append(headers)

    for exp in expenses:
        worksheet.append([
            exp.title,
            exp.amount,
            exp.category.name if exp.category else '',
            exp.date.strftime("%Y-%m-%d"),
            exp.notes or ''
        ])

    for col in worksheet.columns:
        max_length = max(len(str(cell.value)) for cell in col)
        worksheet.column_dimensions[get_column_letter(col[0].column)].width = max_length + 2

    response = HttpResponse(
        content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    )
    response['Content-Disposition'] = 'attachment; filename=expenses.xlsx'
    workbook.save(response)
    return response


# ---------- Export PDF ----------
@login_required
def export_pdf(request):
    expenses = filter_expenses(request, Expense.objects.filter(user=request.user))

    response = HttpResponse(content_type='application/pdf')
    response['Content-Disposition'] = 'attachment; filename="expenses.pdf"'

    doc = SimpleDocTemplate(response, pagesize=A4)
    elements = []
    styles = getSampleStyleSheet()

    elements.append(Paragraph("Expense Report", styles['Title']))
    elements.append(Spacer(1, 12))

    data = [["Title", "Amount", "Category", "Date", "Notes"]]
    for exp in expenses:
        data.append([
            exp.title,
            f"{exp.amount:.2f}",
            exp.category.name if exp.category else '',
            exp.date.strftime("%Y-%m-%d"),
            exp.notes or ''
        ])

    table = Table(data, repeatRows=1)
    table.setStyle(TableStyle([
        ('BACKGROUND', (0,0), (-1,0), colors.grey),
        ('TEXTCOLOR', (0,0), (-1,0), colors.whitesmoke),
        ('ALIGN',(0,0),(-1,-1),'CENTER'),
        ('FONTNAME', (0,0), (-1,0), 'Helvetica-Bold'),
        ('BOTTOMPADDING', (0,0), (-1,0), 8),
        ('BACKGROUND',(0,1),(-1,-1),colors.beige),
        ('GRID',(0,0),(-1,-1),1,colors.black),
    ]))
    elements.append(table)
    doc.build(elements)
    return response


# ---------- Profile ----------
@login_required
def profile(request):
    profile = get_object_or_404(UserProfile, user=request.user)
    if request.method == "POST":
        form = UserProfileForm(request.POST, request.FILES, instance=profile)
        if form.is_valid():
            form.save()
            messages.success(request, "Profile updated!")
            return redirect("profile")
    else:
        form = UserProfileForm(instance=profile)
    return render(request, "tracker/profile.html", {"form": form})


@login_required
def edit_profile(request):
    profile = get_object_or_404(UserProfile, user=request.user)
    if request.method == "POST":
        profile_form = UserProfileForm(request.POST, request.FILES, instance=profile)
        user = request.user
        user.first_name = request.POST.get("first_name")
        user.last_name = request.POST.get("last_name")
        user.email = request.POST.get("email")
        if profile_form.is_valid():
            user.save()
            profile_form.save()
            messages.success(request, "Profile updated successfully!")
            return redirect("profile")
    else:
        profile_form = UserProfileForm(instance=profile)
    return render(request, "tracker/edit_profile.html", {"form": profile_form})


# ---------- Voice Expense ----------
@login_required
@csrf_exempt
def add_expense_voice(request):
    if request.method == "POST":
        try:
            data = json.loads(request.body)
            speech = data.get("speech", "").lower()

            amount_match = re.search(r"\d+(\.\d+)?", speech)
            category_match = None
            categories = Category.objects.all().values_list("name", flat=True)

            for c in categories:
                if c.lower() in speech:
                    category_match = Category.objects.get(name=c)
                    break

            if amount_match:
                amount = float(amount_match.group())
                expense = Expense.objects.create(
                    title=f"Voice Entry - {category_match.name if category_match else 'Other'}",
                    amount=amount,
                    category=category_match,
                    user=request.user
                )
                return JsonResponse({
                    "success": True,
                    "msg": f"✅ Added {amount} in {category_match.name if category_match else 'Other'}"
                })

            return JsonResponse({"success": False, "msg": "⚠️ Could not detect amount."})
        except Exception as e:
            logger.error("Voice Expense Error: %s", str(e))
            return JsonResponse({"success": False, "msg": f"Error: {str(e)}"})
    return JsonResponse({"success": False, "msg": "Invalid request"})


# ---------- AI Chat ----------
@login_required
def ai_chat_page(request):
    return render(request, "tracker/ai_chat.html")

@csrf_exempt
@login_required
def ai_chat(request):
    if request.method == "POST":
        try:
            data = json.loads(request.body)
            user_message = data.get("message", "").strip()
            if not user_message:
                return JsonResponse({"reply": "⚠️ Please type something."})

            # Save user message
            save_message(request.user.id, "user", user_message)

            # Check if we need to summarize
            summarize_old_chats(request.user.id)

            # --- STEP 1: Try to answer from Expense DB ---
            from django.db.models import Sum
            expenses = Expense.objects.filter(user=request.user)

            categories = Category.objects.values_list("name", flat=True)
            matched_category = None
            for c in categories:
                if c.lower() in user_message.lower():
                    matched_category = c
                    break

            # If user asked "how much spent" + mentioned category
            if ("spent" in user_message.lower() or "spend" in user_message.lower()) and matched_category:
                total = expenses.filter(category__name=matched_category).aggregate(Sum("amount"))["amount__sum"] or 0
                reply = f"💰 You have spent a total of {total} on {matched_category}."
            else:
                # --- STEP 2: Fallback to Gemini ---
                summary = load_summary(request.user.id)
                history = load_history(request.user.id)

                history_text = "\n".join([f"{msg['role']}: {msg['content']}" for msg in history])
                prompt = f"User's past memory: {summary}\n\nRecent conversation:\n{history_text}\nuser: {user_message}\nassistant:"

                model = genai.GenerativeModel("gemini-1.5-flash")
                response = model.generate_content(prompt)
                reply = response.text if response and hasattr(response, "text") else "⚠️ No response from AI."

            # Save assistant reply
            save_message(request.user.id, "assistant", reply)

            return JsonResponse({"reply": reply})
        except Exception as e:
            logger.error("AI Chat Error: %s", str(e))
            return JsonResponse({"reply": f"⚠️ Error: {str(e)}"})
    return JsonResponse({"reply": "⚠️ Invalid request method."})

# ---------- Reset AI Memory ----------
@csrf_exempt
@login_required
def reset_ai_memory(request):
    if request.method == "POST":
        try:
            chat_collection.delete_many({"user_id": request.user.id})
            summary_collection.delete_one({"user_id": request.user.id})
            return JsonResponse({"success": True, "msg": "🧹 Memory cleared successfully!"})
        except Exception as e:
            return JsonResponse({"success": False, "msg": f"⚠️ Error: {str(e)}"})
    return JsonResponse({"success": False, "msg": "⚠️ Invalid request method."})
